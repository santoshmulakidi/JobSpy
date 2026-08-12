from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def build_top250(workbook: Path) -> list[dict[str, int | str]]:
    sheet = load_workbook(workbook, read_only=True, data_only=True)["Employer Information"]
    totals: dict[str, dict[str, int]] = defaultdict(lambda: {"new": 0, "continuation": 0})
    for row in sheet.iter_rows(min_row=4, values_only=True):
        name = str(row[2] or "").strip()
        if not name:
            continue
        totals[name]["new"] += int(row[8] or 0)
        totals[name]["continuation"] += int(row[10] or 0)
    ranked = sorted(
        totals.items(),
        key=lambda item: (-(item[1]["new"] + item[1]["continuation"]), item[0].casefold()),
    )[:250]
    return [
        {
            "rank": rank,
            "sponsor_name": name,
            "new_approvals": values["new"],
            "continuation_approvals": values["continuation"],
            "total_approvals": values["new"] + values["continuation"],
        }
        for rank, (name, values) in enumerate(ranked, 1)
    ]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    result = build_top250(args.workbook)
    args.output.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {len(result)} sponsors to {args.output}")


if __name__ == "__main__":
    main()
